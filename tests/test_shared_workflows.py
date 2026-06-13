from __future__ import annotations

import json
import sys
import threading
import time
import types
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
PHOTO_APP = ROOT / "photo_tool" / "app"
if str(PHOTO_APP) not in sys.path:
    sys.path.insert(0, str(PHOTO_APP))

import app
import assignment_engine
import google_sheets_import
from comp_engine.workbook_io import WorkbookRow
from intake_io import append_company_sheet_rows, mark_received_in_workbooks, read_company_profit_records, read_simple_spreadsheet, write_working_sheet
from shared_state import atomic_write_json, local_identity, read_json, shared_lock


if "google" not in sys.modules:
    google_module = types.ModuleType("google")
    genai_module = types.ModuleType("google.genai")
    genai_types_module = types.ModuleType("google.genai.types")

    class _Client:
        pass

    class _Part:
        @staticmethod
        def from_bytes(data=None, mime_type=None):
            return {"data": data, "mime_type": mime_type}

    class _GenerateContentConfig:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

    class _ThinkingConfig:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

    genai_module.Client = _Client
    genai_types_module.Part = _Part
    genai_types_module.GenerateContentConfig = _GenerateContentConfig
    genai_types_module.ThinkingConfig = _ThinkingConfig
    genai_module.types = genai_types_module
    google_module.genai = genai_module
    sys.modules["google"] = google_module
    sys.modules["google.genai"] = genai_module
    sys.modules["google.genai.types"] = genai_types_module

import multi_card_extraction


class SharedStateTests(unittest.TestCase):
    def test_shared_lock_serializes_concurrent_writers(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            events: list[tuple[str, str, float]] = []

            def worker(name: str, delay: float) -> None:
                with shared_lock(root, "same-file", {"display_name": name, "machine": name}):
                    events.append((name, "enter", time.time()))
                    time.sleep(delay)
                    events.append((name, "exit", time.time()))

            first = threading.Thread(target=worker, args=("A", 0.25))
            second = threading.Thread(target=worker, args=("B", 0.01))
            first.start()
            time.sleep(0.05)
            second.start()
            first.join()
            second.join()

            self.assertEqual([event[:2] for event in events], [("A", "enter"), ("A", "exit"), ("B", "enter"), ("B", "exit")])

    def test_atomic_json_write_and_local_identity(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            settings_path = root / "lucas_settings.json"
            identity = local_identity(settings_path)
            self.assertTrue(identity["user_id"])
            self.assertTrue(identity["machine"])

            state_path = root / "state.json"
            atomic_write_json(state_path, {"a": 1})
            atomic_write_json(state_path, {"b": [1, 2, 3]})
            self.assertEqual(read_json(state_path, {}), {"b": [1, 2, 3]})
            self.assertFalse(list(root.glob("*.tmp")))


class WorkbookCompanyProfitTests(unittest.TestCase):
    def test_receive_company_append_dedupes_and_profit_backfills(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            working_dir = root / "WORKING SHEETS"
            company_dir = root / "COMPANY SHEETS"
            working_dir.mkdir()

            source_rows = [
                WorkbookRow(excel_row=2, cert_number="11111111", grader="PSA", card_title="Card One PSA 10", existing_value=50),
                WorkbookRow(excel_row=3, cert_number="22222222", grader="PSA", card_title="Card Two PSA 9", existing_value=75),
            ]
            sheet_path = working_dir / "test.xlsx"
            write_working_sheet(sheet_path, source_rows)

            receive_result = mark_received_in_workbooks([sheet_path], {"11111111"})
            self.assertEqual(receive_result["rows_marked"], 1)
            self.assertIn("11111111", receive_result["certs_marked"])
            self.assertEqual(len(read_simple_spreadsheet(sheet_path)), 2)

            sold_row = WorkbookRow(
                excel_row=2,
                cert_number="11111111",
                grader="PSA",
                card_title="Card One PSA 10",
                existing_value=50,
                card_ladder_value=100,
                card_ladder_comps_average=100,
                best_company="Arena Club",
                estimated_payout=90,
                company_pile=True,
                status="Received",
            )
            first_append = append_company_sheet_rows(company_dir, [sold_row], {2: "test.xlsx:2"}, {2: "test.xlsx"})
            second_append = append_company_sheet_rows(company_dir, [sold_row], {2: "test.xlsx:2"}, {2: "test.xlsx"})

            self.assertEqual(first_append["rows_added"], 1)
            self.assertEqual(second_append["rows_added"], 0)
            self.assertEqual(len(first_append["added_records"]), 1)

            profit_records = read_company_profit_records(company_dir)
            self.assertEqual(len(profit_records), 1)
            self.assertEqual(profit_records[0]["purchase_price"], 50.0)
            self.assertEqual(profit_records[0]["sale_price"], 90.0)


class GoogleSheetCacheTests(unittest.TestCase):
    def test_authenticated_google_sheet_export_writes_xlsx_cache(self) -> None:
        def fake_tabs(_url: str, interactive: bool = False, sheet_name: str = ""):
            return [
                ("Rules/Main*?", [["Category", "Value"], ["Baseball", "100"]]),
                ("Payouts", [["CATEGORY", "YOUR PAYOUT %"], ["Baseball", "90%"]]),
            ]

        with TemporaryDirectory() as tmp, patch.object(google_sheets_import, "read_google_sheet_tabs", side_effect=fake_tabs):
            output_path = Path(tmp) / "cache.xlsx"
            google_sheets_import.export_google_sheet_to_xlsx("https://docs.google.com/spreadsheets/d/test/edit", output_path)

            from openpyxl import load_workbook

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Rules Main", "Payouts"])
                self.assertEqual(workbook["Payouts"].cell(2, 2).value, "90%")
            finally:
                workbook.close()

    def test_startup_google_sheet_cache_refresh_discovers_and_exports_sources(self) -> None:
        class Dummy:
            _saved_google_sheet_sources = app.CardPipelineApp._saved_google_sheet_sources
            _google_sheet_cache_source = app.CardPipelineApp._google_sheet_cache_source
            _refresh_startup_google_sheet_caches = app.CardPipelineApp._refresh_startup_google_sheet_caches

        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_config = app.ASSIGNMENT_CONFIG_PATH
            app.CARD_PIPELINE_DIR = tmp_path / "CARD_PIPELINE"
            app.ASSIGNMENT_CONFIG_PATH = tmp_path / "assignment_companies.json"
            cache_path = app.CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "SHEET EXPORTS" / "rules.xlsx"
            app.ASSIGNMENT_CONFIG_PATH.write_text(
                json.dumps(
                    {
                        "companies": [
                            {
                                "name": "Test",
                                "rules": {
                                    "kind": "google_sheet",
                                    "url": "https://docs.google.com/spreadsheets/d/abc/edit",
                                    "path": str(cache_path),
                                    "name": "Rules",
                                },
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            calls: list[tuple[str, Path, bool]] = []

            def fake_export(url: str, output_path: Path, interactive: bool = False) -> Path:
                calls.append((url, Path(output_path), interactive))
                Path(output_path).parent.mkdir(parents=True, exist_ok=True)
                Path(output_path).write_bytes(b"fake")
                return Path(output_path)

            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                with patch.object(app, "export_google_sheet_to_xlsx", side_effect=fake_export):
                    result = dummy._refresh_startup_google_sheet_caches()
                self.assertEqual(result["refreshed"], 1)
                self.assertEqual(result["errors"], [])
                self.assertEqual(calls[0][0], "https://docs.google.com/spreadsheets/d/abc/edit")
                self.assertFalse(calls[0][2])
                self.assertTrue(cache_path.exists())
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.ASSIGNMENT_CONFIG_PATH = old_config


class AssignmentEngineTests(unittest.TestCase):
    def test_recommendation_chooses_highest_payout_among_accepted_companies(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="1",
            grader="PSA",
            card_title="2019 Panini Prizm Stephen Curry Silver PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Lower",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NBA")],
                ),
                assignment_engine.AssignmentCompany(
                    "Higher",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.95, "NBA")],
                ),
                assignment_engine.AssignmentCompany(
                    "Rejected",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("baseball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 1.0, "MLB")],
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Higher")
        self.assertEqual(recommendation.payout, 95)
        self.assertTrue(decisions["Lower"].accepted)
        self.assertFalse(decisions["Rejected"].accepted)

    def test_company_can_prefer_card_ladder_value_over_comps(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="4",
            grader="PSA",
            card_title="2020 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Comps Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.95, "NFL")],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "CL Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NFL")],
                    value_source="card_ladder",
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(decisions["Comps Buyer"].source_value, 100)
        self.assertEqual(decisions["Comps Buyer"].payout, 95)
        self.assertEqual(decisions["CL Buyer"].source_value, 150)
        self.assertEqual(decisions["CL Buyer"].payout, 135)
        self.assertEqual(recommendation.company, "CL Buyer")
        self.assertEqual(recommendation.payout, 135)

    def test_card_ladder_value_source_rejects_company_when_cl_missing(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="5",
            grader="PSA",
            card_title="2020 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=None,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Comps Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NFL")],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "CL Required Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 1.0, "NFL")],
                    value_source="card_ladder",
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Comps Buyer")
        self.assertEqual(recommendation.payout, 90)
        self.assertFalse(decisions["CL Required Buyer"].accepted)
        self.assertIsNone(decisions["CL Required Buyer"].source_value)
        self.assertIn("missing", decisions["CL Required Buyer"].reason)

    def test_goat_payout_category_uses_payout_range_not_rule_goat_range(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="2",
            grader="PSA",
            card_title="2019 Panini Mosaic Stephen Curry Green Mosaic PSA 10",
            card_ladder_comps_average=85.61,
        )
        rules = assignment_engine.CompanyRules(
            ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)],
            goat_players={"stephen curry"},
            goat_ranges=[assignment_engine.AssignmentRule("Stephen Curry", 100, 7500)],
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Goat Buyer",
                    rules,
                    [assignment_engine.PayoutTier(50, 99, 0.95, "GOATS")],
                )
            ]
        )

        recommendation = engine.recommend(row)

        self.assertEqual(recommendation.company, "Goat Buyer")
        self.assertEqual(recommendation.payout, 81.33)

    def test_accepted_company_without_matching_payout_cannot_win(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="3",
            grader="PSA",
            card_title="2022 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=80,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "No Payout Match",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(100, 500, 1.0, "NFL")],
                ),
                assignment_engine.AssignmentCompany(
                    "Valid Payout",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 99, 0.9, "NFL")],
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Valid Payout")
        self.assertIsNone(decisions["No Payout Match"].payout)
        self.assertIn("no payout tier", decisions["No Payout Match"].reason)


class AppSharedWorkflowLogicTests(unittest.TestCase):
    def test_sheet_marker_save_merges_latest_and_honors_tombstones(self) -> None:
        class MarkerDummy:
            _load_sheet_markers = app.CardPipelineApp._load_sheet_markers
            _save_sheet_markers = app.CardPipelineApp._save_sheet_markers
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_markers = app.SHEET_MARKERS_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.SHEET_MARKERS_PATH = Path(tmp) / "sheet_markers.json"
            app.SHEET_MARKERS_PATH.write_text(
                json.dumps(
                    {
                        "Incoming|A.xlsx": {"assigned_person": "A"},
                        "Incoming|B.xlsx": {"assigned_person": "B"},
                    }
                ),
                encoding="utf-8",
            )
            dummy = MarkerDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_sheet_markers = {"Incoming|C.xlsx": {"assigned_person": "C"}}
            dummy.deleted_sheet_marker_keys = {"Incoming|A.xlsx"}
            try:
                dummy._save_sheet_markers()
                saved = json.loads(app.SHEET_MARKERS_PATH.read_text(encoding="utf-8"))
                self.assertNotIn("Incoming|A.xlsx", saved)
                self.assertEqual(saved["Incoming|B.xlsx"]["assigned_person"], "B")
                self.assertEqual(saved["Incoming|C.xlsx"]["assigned_person"], "C")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.SHEET_MARKERS_PATH = old_markers

    def test_profit_sales_are_deduped_and_delta_is_recorded(self) -> None:
        class ProfitDummy:
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_profit_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            dummy = ProfitDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                record = {
                    "date_added": "2026-06-11",
                    "company": "Arena Club",
                    "weekly_sheet_name": "Arena WEEK.xlsx",
                    "source_sheet": "source.xlsx",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "purchase_price": "$40.00",
                    "sale_price": "$90.00",
                }
                self.assertEqual(dummy.record_profit_sales([record]), 1)
                self.assertEqual(dummy.record_profit_sales([record]), 0)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["profit"], 50.0)
                self.assertEqual(ledger[0]["recorded_by"], "Tester")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_ledger


class PhotoOcrSpeedTests(unittest.TestCase):
    def test_detect_regions_skips_extra_label_sweeps_when_dense_target_is_met(self) -> None:
        calls: list[str] = []

        def fake_detect(_client, _bytes, _mime, prompt):
            calls.append(prompt)
            if prompt == multi_card_extraction.DETECTION_PROMPT:
                return [
                    {"card_index": index + 1, "position": f"slot {index + 1}", "bbox": [index * 50, 0, index * 50 + 40, 400], "detection_confidence": "high"}
                    for index in range(multi_card_extraction.PHOTO_OCR_REGION_TARGET)
                ]
            return []

        with patch.object(multi_card_extraction, "_detect_regions_for_prompt", side_effect=fake_detect), \
                patch.object(multi_card_extraction, "_detect_best_row_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_add_uncovered_edge_regions", side_effect=lambda regions: regions):
            regions = multi_card_extraction._detect_regions_sync(object(), b"image", "image/jpeg")

        self.assertEqual(len(regions), multi_card_extraction.PHOTO_OCR_REGION_TARGET)
        self.assertIn(multi_card_extraction.DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_SWEEP_PROMPT, calls)

    def test_detect_regions_uses_label_sweeps_below_dense_target(self) -> None:
        calls: list[str] = []

        def fake_detect(_client, _bytes, _mime, prompt):
            calls.append(prompt)
            if prompt == multi_card_extraction.DETECTION_PROMPT:
                return [
                    {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
                    {"card_index": 2, "position": "middle", "bbox": [220, 0, 420, 400], "detection_confidence": "high"},
                    {"card_index": 3, "position": "right", "bbox": [440, 0, 640, 400], "detection_confidence": "high"},
                ]
            if prompt == multi_card_extraction.LABEL_DETECTION_PROMPT:
                return [
                    {"card_index": index + 1, "position": f"label {index + 1}", "bbox": [index * 50, 0, index * 50 + 40, 100], "detection_confidence": "high"}
                    for index in range(6)
                ]
            return []

        with patch.object(multi_card_extraction, "_detect_regions_for_prompt", side_effect=fake_detect), \
                patch.object(multi_card_extraction, "_detect_best_row_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_detect_best_prompt_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_add_uncovered_edge_regions", side_effect=lambda regions: regions):
            regions = multi_card_extraction._detect_regions_sync(object(), b"image", "image/jpeg")

        self.assertTrue(regions)
        self.assertIn(multi_card_extraction.LABEL_DETECTION_PROMPT, calls)

    def test_identify_cards_reports_crop_progress_and_preserves_order(self) -> None:
        callbacks: list[str] = []
        regions = [
            {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
            {"card_index": 2, "position": "right", "bbox": [220, 0, 420, 400], "detection_confidence": "medium"},
        ]

        def fake_identify(_client, crop_b64):
            return {
                "is_graded_slab": True,
                "grading_company": "PSA",
                "cert_number": "111" if crop_b64 == "crop-1" else "222",
                "player": "Player",
                "year": "2020",
                "set": "Test",
                "card_number": "",
                "parallel": "",
                "subset": "",
                "grade": "10",
                "category": "baseball",
                "confidence": "high",
                "label_text": "label",
            }

        with patch.object(multi_card_extraction, "_prepare_image", return_value=(b"image", "image/jpeg")), \
                patch.object(multi_card_extraction, "_detect_regions_sync", return_value=regions), \
                patch.object(multi_card_extraction, "_decode_image", return_value=object()), \
                patch.object(multi_card_extraction, "_crop_region_to_base64", side_effect=["crop-1", "crop-2"]), \
                patch.object(multi_card_extraction, "_identify_crop_sync", side_effect=fake_identify):
            cards = multi_card_extraction.identify_cards_sync(object(), "fake-b64", progress_callback=callbacks.append)

        self.assertEqual([card["cert_number"] for card in cards], ["111", "222"])
        self.assertTrue(any("Detected 2 card(s)" in message for message in callbacks))
        self.assertTrue(any("Read 2/2" in message for message in callbacks))

    def test_identify_cards_keeps_detected_slab_when_crop_ocr_fails(self) -> None:
        regions = [
            {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
            {"card_index": 2, "position": "right", "bbox": [220, 0, 420, 400], "detection_confidence": "medium"},
        ]

        def fake_identify(_client, crop_b64):
            if crop_b64 == "crop-2":
                raise RuntimeError("label unreadable")
            return {
                "is_graded_slab": True,
                "grading_company": "PSA",
                "cert_number": "111",
                "player": "Player",
                "year": "2020",
                "set": "Test",
                "card_number": "",
                "parallel": "",
                "subset": "",
                "grade": "10",
                "category": "baseball",
                "confidence": "high",
                "label_text": "label",
            }

        with patch.object(multi_card_extraction, "_prepare_image", return_value=(b"image", "image/jpeg")), \
                patch.object(multi_card_extraction, "_detect_regions_sync", return_value=regions), \
                patch.object(multi_card_extraction, "_decode_image", return_value=object()), \
                patch.object(multi_card_extraction, "_crop_region_to_base64", side_effect=["crop-1", "crop-2"]), \
                patch.object(multi_card_extraction, "_identify_crop_sync", side_effect=fake_identify):
            cards = multi_card_extraction.identify_cards_sync(object(), "fake-b64")

        self.assertEqual(len(cards), 2)
        self.assertEqual(cards[0]["cert_number"], "111")
        self.assertEqual(cards[1]["card_index"], 2)
        self.assertTrue(cards[1]["is_graded_slab"])
        self.assertIn("label unreadable", cards[1]["error"])

    def test_photo_table_accepts_detected_slab_without_readable_inventory(self) -> None:
        card = {
            "card_index": 2,
            "position": "right",
            "is_graded_slab": True,
            "detection_confidence": "medium",
            "error": "label unreadable",
        }

        self.assertTrue(app.CardPipelineApp._photo_card_has_inventory(object(), card))
        row = app.CardPipelineApp._photo_card_to_row(object(), Path("dense.jpg"), card)
        self.assertEqual(row["source"], "Photo: dense.jpg")
        self.assertIn("right", row["notes"])
        self.assertIn("OCR review needed", row["notes"])


if __name__ == "__main__":
    unittest.main()
