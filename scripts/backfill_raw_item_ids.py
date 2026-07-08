from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


RAW_ID_PATTERN = re.compile(r"^RAW-\d{8}-\d{4}$", re.IGNORECASE)
ITEM_ID_HEADER = "Item ID"
ITEM_ID_ALIASES = {"itemid", "rawitemid", "inventoryitemid", "inventoryid"}
CERT_ALIASES = {"certificationnumber", "certnumber", "cert", "certification", "cert#"}
CARD_ALIASES = {"carddescription", "card", "description", "title", "cardtitle", "item", "itemtitle"}


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9#]", "", str(value or "").strip().lower())


def _headers(sheet) -> dict[str, int]:
    return {
        _normalize_header(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if str(cell.value or "").strip()
    }


def _first_header_col(headers: dict[str, int], aliases: set[str]) -> int | None:
    for alias in aliases:
        column = headers.get(alias)
        if column:
            return column
    return None


def _looks_like_header(sheet) -> bool:
    headers = _headers(sheet)
    return bool(_first_header_col(headers, CARD_ALIASES) or _first_header_col(headers, CERT_ALIASES))


def _next_raw_id(existing: set[str], sequence_date: str) -> str:
    prefix = f"RAW-{sequence_date}-"
    max_sequence = 0
    for value in existing:
        text = str(value or "").strip().upper()
        if text.startswith(prefix):
            suffix = text[len(prefix):]
            if suffix.isdigit():
                max_sequence = max(max_sequence, int(suffix))
    while True:
        max_sequence += 1
        candidate = f"{prefix}{max_sequence:04d}"
        if candidate not in existing:
            existing.add(candidate)
            return candidate


def _collect_existing_ids(root: Path) -> set[str]:
    existing: set[str] = set()
    ledger_path = root / "inventory_ledger.json"
    if ledger_path.exists():
        try:
            payload = json.loads(ledger_path.read_text(encoding="utf-8"))
            for item in payload.get("items") or []:
                item_id = str(item.get("item_id") or "").strip().upper()
                if item_id:
                    existing.add(item_id)
        except Exception:
            pass
    for folder_name in ("INCOMING SHEETS", "WORKING SHEETS", "RECEIVED SHEETS"):
        folder = root / folder_name
        if not folder.exists():
            continue
        for path in sorted(folder.glob("*.xlsx"), key=lambda item: item.name.lower()):
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            try:
                for sheet in workbook.worksheets:
                    if not _looks_like_header(sheet):
                        continue
                    headers = _headers(sheet)
                    item_id_col = _first_header_col(headers, ITEM_ID_ALIASES)
                    if not item_id_col:
                        continue
                    for row_index in range(2, sheet.max_row + 1):
                        item_id = str(sheet.cell(row_index, item_id_col).value or "").strip().upper()
                        if item_id:
                            existing.add(item_id)
            finally:
                workbook.close()
    return existing


def backfill_root(root: Path, dry_run: bool = False) -> dict[str, object]:
    existing_ids = _collect_existing_ids(root)
    sequence_date = datetime.now().strftime("%Y%m%d")
    summary = {"root": str(root), "files_changed": 0, "ids_added": 0, "files": []}
    for folder_name in ("INCOMING SHEETS", "WORKING SHEETS"):
        folder = root / folder_name
        if not folder.exists():
            continue
        for path in sorted(folder.glob("*.xlsx"), key=lambda item: item.name.lower()):
            workbook = load_workbook(path)
            changed = False
            added_for_file = 0
            try:
                for sheet in workbook.worksheets:
                    if not _looks_like_header(sheet):
                        continue
                    headers = _headers(sheet)
                    cert_col = _first_header_col(headers, CERT_ALIASES)
                    card_col = _first_header_col(headers, CARD_ALIASES)
                    if not card_col:
                        continue
                    item_id_col = _first_header_col(headers, ITEM_ID_ALIASES)
                    if not item_id_col:
                        item_id_col = sheet.max_column + 1
                        sheet.cell(1, item_id_col).value = ITEM_ID_HEADER
                        changed = True
                    for row_index in range(2, sheet.max_row + 1):
                        cert = str(sheet.cell(row_index, cert_col).value or "").strip() if cert_col else ""
                        card = str(sheet.cell(row_index, card_col).value or "").strip()
                        item_id = str(sheet.cell(row_index, item_id_col).value or "").strip()
                        if cert or not card or item_id:
                            continue
                        sheet.cell(row_index, item_id_col).value = _next_raw_id(existing_ids, sequence_date)
                        changed = True
                        added_for_file += 1
                if changed and not dry_run:
                    workbook.save(path)
            finally:
                workbook.close()
            if added_for_file:
                summary["files_changed"] += 1
                summary["ids_added"] += added_for_file
                summary["files"].append({"path": str(path), "ids_added": added_for_file})
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill stable Item ID values for raw incoming/working LUCAS rows.")
    parser.add_argument("roots", nargs="+", help="LUCAS root folder(s), such as G:\\My Drive\\CARD_PIPELINE")
    parser.add_argument("--dry-run", action="store_true", help="Report changes without saving workbooks.")
    args = parser.parse_args()
    summaries = [backfill_root(Path(root), dry_run=args.dry_run) for root in args.roots]
    print(json.dumps(summaries, indent=2))


if __name__ == "__main__":
    main()
