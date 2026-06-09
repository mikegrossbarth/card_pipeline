from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\User\Downloads\Blez x mikey MASTER SHEET.xlsx")
DEFAULT_SHEET = "612026"


@dataclass
class WorkbookRow:
    excel_row: int
    cert_number: str
    card_title: str
    grader: str
    existing_value: Any = None
    card_ladder_value: float | None = None
    card_ladder_comps_average: float | None = None
    card_ladder_comp_confidence: str = ""
    card_ladder_comps: str = ""
    card_ladder_screenshot: str = ""
    alt_value: float | None = None
    cy_value: float | None = None
    best_company: str = ""
    estimated_payout: float | None = None
    status: str = "Ready"
    notes: str = ""


def workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_card_rows(path: Path, sheet_name: str, force: bool = False) -> list[WorkbookRow]:
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        headers = _header_map(sheet)
        cert_col = _required_any(headers, "certificationnumber", "certnumber", "cert")
        card_col = _required_any(headers, "card", "carddescription", "description")
        value_col = _optional_any(headers, "value", "cardladdervalue")

        rows: list[WorkbookRow] = []
        for row_index in range(2, sheet.max_row + 1):
            cert = normalize_cert(sheet.cell(row_index, cert_col).value)
            card = str(sheet.cell(row_index, card_col).value or "").strip()
            current_value = sheet.cell(row_index, value_col).value if value_col else None
            if not cert and not card:
                continue
            if not force and not _is_blank(current_value):
                continue
            grader = infer_grader(card)
            status = "Ready" if cert and card and grader else "Needs setup"
            notes = ""
            if not cert:
                notes = "Missing cert"
            elif not grader:
                notes = "Missing grader"
            rows.append(
                WorkbookRow(
                    excel_row=row_index,
                    cert_number=cert,
                    card_title=card,
                    grader=grader,
                    existing_value=current_value,
                    status=status,
                    notes=notes,
                )
            )
        return rows
    finally:
        workbook.close()


def write_results(path: Path, sheet_name: str, rows: list[WorkbookRow], output_path: Path) -> Path:
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        headers = _header_map(sheet)
        value_col = ensure_header(sheet, "Value")
        extra_cols = {
            "cardladdervalue": ensure_header(sheet, "Card Ladder Value"),
            "comps": ensure_header(sheet, "Comps"),
            "compconfidence": ensure_header(sheet, "Comp Confidence"),
            "cardladdercompdetails": ensure_header(sheet, "Card Ladder Comp Details"),
            "cardladderscreenshot": ensure_header(sheet, "Card Ladder Screenshot"),
            "altvalue": ensure_header(sheet, "ALT Value"),
            "cyvalue": ensure_header(sheet, "CY Value"),
            "bestcompany": ensure_header(sheet, "Best Company"),
            "estimatedpayout": ensure_header(sheet, "Estimated Payout"),
            "compstatus": ensure_header(sheet, "Comp Status"),
        }
        by_row = {row.excel_row: row for row in rows}
        for excel_row, row in by_row.items():
            if row.card_ladder_value is not None:
                sheet.cell(excel_row, value_col).value = row.card_ladder_value
                sheet.cell(excel_row, extra_cols["cardladdervalue"]).value = row.card_ladder_value
            if row.card_ladder_comps_average is not None:
                sheet.cell(excel_row, extra_cols["comps"]).value = row.card_ladder_comps_average
            if row.card_ladder_comp_confidence:
                sheet.cell(excel_row, extra_cols["compconfidence"]).value = row.card_ladder_comp_confidence
            if row.card_ladder_comps:
                sheet.cell(excel_row, extra_cols["cardladdercompdetails"]).value = row.card_ladder_comps
            if row.card_ladder_screenshot:
                sheet.cell(excel_row, extra_cols["cardladderscreenshot"]).value = row.card_ladder_screenshot
            if row.alt_value is not None:
                sheet.cell(excel_row, extra_cols["altvalue"]).value = row.alt_value
            if row.cy_value is not None:
                sheet.cell(excel_row, extra_cols["cyvalue"]).value = row.cy_value
            sheet.cell(excel_row, extra_cols["bestcompany"]).value = row.best_company
            if row.estimated_payout is not None:
                sheet.cell(excel_row, extra_cols["estimatedpayout"]).value = row.estimated_payout
            sheet.cell(excel_row, extra_cols["compstatus"]).value = row.status
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
    finally:
        workbook.close()


def ensure_header(sheet, header: str) -> int:
    headers = _header_map(sheet)
    key = _normalize_header(header)
    if key in headers:
        return headers[key]
    col = sheet.max_column + 1
    sheet.cell(1, col).value = header
    return col


def infer_grader(card_title: str) -> str:
    match = re.search(r"\b(PSA|BGS|SGC|CGC)\b", str(card_title or ""), re.I)
    return match.group(1).upper() if match else ""


def normalize_cert(value: Any) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"[^0-9A-Z]", "", str(value or ""), flags=re.I)


def _header_map(sheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if value:
            result[_normalize_header(value)] = col
    return result


def _required(headers: dict[str, int], key: str) -> int:
    if key not in headers:
        raise ValueError(f"Missing required column: {key}")
    return headers[key]


def _required_any(headers: dict[str, int], *keys: str) -> int:
    for key in keys:
        if key in headers:
            return headers[key]
    raise ValueError(f"Missing required column: {' / '.join(keys)}")


def _optional_any(headers: dict[str, int], *keys: str) -> int | None:
    for key in keys:
        if key in headers:
            return headers[key]
    return None


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""
