from __future__ import annotations

import csv
import json
import os
import re
import urllib.parse
import urllib.request
import urllib.error
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from google_sheets_import import GoogleSheetsAuthError, read_google_sheet_text


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "assignment_companies.json"


@dataclass
class AssignmentRule:
    matcher: str = ""
    min_price: float | None = None
    max_price: float | None = None
    block: bool = False


@dataclass
class GradeRule:
    allowed: bool = True
    min_grade: float | None = None
    max_grade: float | None = None


@dataclass
class CompanyRules:
    include: list[str] = field(default_factory=list)
    exclude: list[str] = field(default_factory=list)
    ranges: list[AssignmentRule] = field(default_factory=list)
    blocks: list[AssignmentRule] = field(default_factory=list)
    grade_rules: dict[str, GradeRule] = field(default_factory=dict)
    rule_groups: list["CompanyRules"] = field(default_factory=list)
    accept_all: bool = False


@dataclass
class PayoutTier:
    min_price: float = 0
    max_price: float | None = None
    rate: float = 0


@dataclass
class AssignmentCompany:
    name: str
    rules: CompanyRules
    payout_tiers: list[PayoutTier]


@dataclass
class AssignmentRecommendation:
    company: str = ""
    payout: float | None = None
    source_value: float | None = None


class AssignmentEngine:
    def __init__(self, companies: list[AssignmentCompany] | None = None, error: str = "") -> None:
        self.companies = companies or []
        self.error = error

    @classmethod
    def load(cls, config_path: Path | None = None) -> "AssignmentEngine":
        path = Path(config_path or CONFIG_PATH)
        if not path.exists():
            return cls([])
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            entries = raw.get("companies", raw) if isinstance(raw, dict) else raw
            companies = [load_company(entry, path.parent) for entry in entries if isinstance(entry, dict)]
            return cls([company for company in companies if company is not None])
        except Exception as error:
            return cls([], str(error))

    def recommend(self, row: Any) -> AssignmentRecommendation:
        source_value = assignment_value(row)
        if source_value is None:
            return AssignmentRecommendation()

        card_text = card_row_text(row, source_value)
        candidates: list[AssignmentRecommendation] = []
        for company in self.companies:
            if not company_accepts(company.rules, card_text, source_value, str(getattr(row, "grader", "") or "")):
                continue
            payout = payout_for_value(company.payout_tiers, source_value)
            if payout is None:
                continue
            candidates.append(AssignmentRecommendation(company.name, round(payout, 2), source_value))
        return max(candidates, key=lambda item: item.payout or 0) if candidates else AssignmentRecommendation(source_value=source_value)


def load_company(entry: dict[str, Any], base_dir: Path) -> AssignmentCompany | None:
    name = str(entry.get("name") or "").strip()
    if not name:
        return None
    rules_text = read_source_text(entry.get("rules") or entry.get("rules_source") or entry.get("rulesSource"), base_dir)
    payout_text = read_source_text(entry.get("payout") or entry.get("payout_source") or entry.get("payoutSource"), base_dir)
    rules = parse_rules(rules_text, accept_all=bool(entry.get("accept_all")))
    payout_tiers = parse_payouts(payout_text)
    if not payout_tiers and entry.get("rate") is not None:
        rate = parse_rate(entry.get("rate"))
        if rate is not None:
            payout_tiers = [PayoutTier(rate=rate)]
    return AssignmentCompany(name=name, rules=rules, payout_tiers=payout_tiers)


def assignment_value(row: Any) -> float | None:
    comps = to_number(getattr(row, "card_ladder_comps_average", None))
    cl_value = to_number(getattr(row, "card_ladder_value", None))
    return comps if comps is not None else cl_value


def card_row_text(row: Any, source_value: float) -> str:
    parts = [
        getattr(row, "cert_number", ""),
        getattr(row, "grader", ""),
        getattr(row, "card_title", ""),
        f"${source_value}",
    ]
    return " ".join(str(part or "") for part in parts).strip()


def company_accepts(rules: CompanyRules, text: str, price: float, grader: str) -> bool:
    if not rules.accept_all and not (rules.include or rules.exclude or rules.ranges or rules.blocks or rules.grade_rules or rules.rule_groups):
        return False
    haystack = clean_text(text)
    grade_company, grade = parse_grade(text, grader)

    for rule in rules.blocks:
        if rule_matches(rule, haystack, price):
            return False
    if any(term_matches(term, haystack) for term in rules.exclude):
        return False
    if rules.rule_groups:
        return any(company_accepts(group, text, price, grader) for group in rules.rule_groups)

    if rules.grade_rules:
        grade_rule = rules.grade_rules.get(clean_text(grade_company))
        if grade_rule is None or not grade_rule.allowed:
            return False
        if grade is not None and grade_rule.min_grade is not None and grade < grade_rule.min_grade:
            return False
        if grade is not None and grade_rule.max_grade is not None and grade > grade_rule.max_grade:
            return False

    if rules.include and not any(term_matches(term, haystack) for term in rules.include):
        return False
    if rules.ranges:
        return any(rule_matches(rule, haystack, price) for rule in rules.ranges)
    return True


def payout_for_value(tiers: list[PayoutTier], value: float) -> float | None:
    for tier in tiers:
        if value < tier.min_price:
            continue
        if tier.max_price is not None and value > tier.max_price:
            continue
        return value * tier.rate
    return None


def read_source_text(source: Any, base_dir: Path, interactive_google: bool = False) -> str:
    if isinstance(source, dict):
        return read_structured_source_text(source, base_dir, interactive_google=interactive_google)
    raw = normalize_source_value(source)
    if not raw:
        return ""
    if raw.startswith(("http://", "https://")):
        return read_url_text(raw)
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = base_dir / path
    try:
        if not path.exists():
            return ""
    except OSError as error:
        raise ValueError(f"Invalid local source path: {raw}") from error
    if path.suffix.lower() == ".gsheet":
        exported = materialize_gsheet_shortcut(path, path.parent / "LUCAS SHEET EXPORTS")
        return read_workbook_text(exported)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_workbook_text(path)
    try:
        return path.read_text(encoding="utf-8-sig")
    except OSError as error:
        raise ValueError(f"Could not open local source path: {raw}") from error


def read_structured_source_text(source: dict[str, Any], base_dir: Path, interactive_google: bool = False) -> str:
    kind = str(source.get("kind") or "").strip()
    path_value = source.get("path") or source.get("file")
    url = str(source.get("url") or "").strip()
    if kind == "google_sheet" and url:
        path = path_from_source_value(path_value, base_dir) if path_value else None
        try:
            return read_google_sheet_text(url, interactive=interactive_google)
        except GoogleSheetsAuthError:
            if not path or not path.exists():
                raise
        except Exception:
            if not path or not path.exists():
                raise
        try:
            if path:
                materialize_google_sheet_url_to_path(url, path)
                return read_workbook_text(path)
            output_dir = base_dir / "ASSIGNMENT RULES" / "SHEET EXPORTS"
            exported = materialize_google_sheet_url(url, output_dir, str(source.get("name") or "google-sheet"), unique=False)
            return read_workbook_text(exported)
        except Exception:
            if path and path.exists():
                return read_workbook_text(path)
            raise
    return read_source_text(path_value or url or source.get("doc_id"), base_dir, interactive_google=interactive_google)


def path_from_source_value(value: Any, base_dir: Path) -> Path:
    raw = normalize_source_value(value)
    path = Path(raw).expanduser()
    return path if path.is_absolute() else base_dir / path


def normalize_source_value(source: Any) -> str:
    raw = str(source or "").strip().strip('"').strip("'")
    if raw.startswith("file://"):
        parsed = urllib.parse.urlparse(raw)
        raw = urllib.parse.unquote(parsed.path or "")
        if os.name == "nt" and re.match(r"^/[A-Za-z]:/", raw):
            raw = raw[1:]
    if os.name == "nt" and re.match(r"^/[A-Za-z]:[\\/]", raw):
        raw = raw[1:]
    return raw


def load_gsheet_shortcut(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_bytes()
    except OSError as error:
        raise ValueError(
            f"Could not open this .gsheet shortcut locally: {path}. Google Drive may be exposing it as an unreadable placeholder."
        ) from error
    try:
        shortcut = json.loads(raw.decode("utf-8-sig"))
    except UnicodeDecodeError as error:
        raise ValueError(
            f"This .gsheet shortcut is not UTF-8 JSON: {path}. Paste the Google Sheet URL so L.U.C.A.S can export it."
        ) from error
    except json.JSONDecodeError as error:
        raise ValueError(
            f"This .gsheet file is not readable shortcut JSON: {path}. Paste the Google Sheet URL so L.U.C.A.S can export it."
        ) from error
    return shortcut


def read_gsheet_shortcut_text(path: Path) -> str:
    shortcut = load_gsheet_shortcut(path)
    url = gsheet_shortcut_url(shortcut)
    if not url:
        raise ValueError(
            "This .gsheet shortcut does not contain readable sheet data. Choose a synced/exported .xlsx or .csv copy from Google Drive."
        )
    text = read_url_text(url)
    if not text.strip():
        raise ValueError(
            "Google returned no CSV rows for this .gsheet shortcut. Export or sync the sheet as .xlsx/.csv and choose that file."
        )
    return text


def materialize_gsheet_shortcut(path: Path, output_dir: Path) -> Path:
    shortcut = load_gsheet_shortcut(path)
    url = gsheet_shortcut_url(shortcut)
    if not url:
        raise ValueError("This .gsheet shortcut does not contain a Google Sheet URL or document id.")
    return materialize_google_sheet_url(url, output_dir, str(shortcut.get("name") or path.stem or "google-sheet"))


def materialize_google_sheet_url(url: str, output_dir: Path, name: str = "google-sheet", unique: bool = True) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(name)
    output_path = output_dir / f"{stem}.xlsx"
    if unique:
        output_path = unique_export_path(output_path)
    materialize_google_sheet_url_to_path(url, output_path)
    return output_path


def materialize_google_sheet_url_to_path(url: str, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    download_google_sheet_xlsx(url, output_path)


def download_google_sheet_xlsx(url: str, output_path: Path) -> None:
    request_url = google_sheet_xlsx_url(url)
    try:
        with urllib.request.urlopen(request_url, timeout=40) as response:
            data = response.read()
    except urllib.error.HTTPError as error:
        if error.code in {401, 403}:
            raise ValueError(
                "Google rejected the sheet export because this sheet is private. Share/export access is required, or L.U.C.A.S needs an authenticated Google import connection."
            ) from error
        raise
    if not data.startswith(b"PK"):
        text = data[:500].decode("utf-8", errors="replace")
        if re.search(r"<!doctype html|<html[\s>]", text, re.I):
            raise ValueError(
                "Google returned a web page instead of an XLSX export. Open the sheet in Google Drive and save/download it as .xlsx or .csv."
            )
        raise ValueError("Google did not return a valid XLSX export for this sheet.")
    output_path.write_bytes(data)


def google_sheet_xlsx_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return url
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return url
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"


def unique_export_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}-{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    return path


def safe_filename(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._ -]+", "-", value).strip(" .-")
    return safe or "google-sheet"


def gsheet_shortcut_url(shortcut: dict[str, Any]) -> str:
    url = str(shortcut.get("url") or "").strip()
    if url:
        return url
    doc_id = str(shortcut.get("doc_id") or "").strip()
    resource_id = str(shortcut.get("resource_id") or "").strip()
    if not doc_id and resource_id.startswith("spreadsheet:"):
        doc_id = resource_id.split(":", 1)[1].strip()
    if doc_id:
        return f"https://docs.google.com/spreadsheets/d/{doc_id}/edit"
    return ""


def read_url_text(url: str) -> str:
    request_url = google_sheet_csv_url(url) or url
    with urllib.request.urlopen(request_url, timeout=20) as response:
        data = response.read()
    text = data.decode("utf-8-sig", errors="replace")
    if re.search(r"^\s*<!doctype html|<html[\s>]", text, re.I):
        raise ValueError(
            "Google returned a web page instead of sheet rows. Choose a synced/exported .xlsx or .csv file from your Drive folder."
        )
    return text


def google_sheet_csv_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return ""
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return ""
    query = urllib.parse.parse_qs(parsed.query)
    gid = query.get("gid", ["0"])[0]
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv&gid={gid}"


def read_workbook_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                if cells:
                    lines.append(" ".join(cells))
        return "\n".join(lines)
    finally:
        workbook.close()


def parse_rules(text: str, accept_all: bool = False) -> CompanyRules:
    stripped = str(text or "").strip()
    if not stripped:
        return CompanyRules(accept_all=accept_all)
    try:
        payload = json.loads(stripped)
        if isinstance(payload, dict):
            return parse_rule_dict(payload, accept_all)
    except Exception:
        pass

    rules = CompanyRules(accept_all=accept_all)
    for line in source_lines(stripped):
        lowered = line.lower()
        key_value = re.match(r"^([^:=]+)\s*[:=]\s*(.+)$", line)
        if key_value:
            key = normalize_key(key_value.group(1))
            value = key_value.group(2).strip()
            if key in {"include", "includekeywords", "keywords", "sports", "sport"}:
                rules.include.extend(split_values(value))
                continue
            if key in {"exclude", "excludekeywords"}:
                rules.exclude.extend(split_values(value))
                continue
            if key in {"block", "blocks", "blockrules", "donotbuy", "neverbuy"}:
                rules.blocks.append(parse_rule_line(value, block=True))
                continue
            if key in {"minprice", "minimumprice"}:
                rules.ranges.append(AssignmentRule(min_price=parse_money(value)))
                continue
            if key in {"maxprice", "maximumprice"}:
                rules.ranges.append(AssignmentRule(max_price=parse_money(value)))
                continue
        if "block:" in lowered:
            rules.blocks.append(parse_rule_line(line.split(":", 1)[1], block=True))
            continue
        parsed = parse_rule_line(line)
        if parsed.min_price is not None or parsed.max_price is not None:
            rules.ranges.append(parsed)
    return rules


def parse_rule_dict(payload: dict[str, Any], accept_all: bool = False) -> CompanyRules:
    rules = CompanyRules(accept_all=accept_all or bool(payload.get("accept_all")))
    rules.include.extend(split_values(payload.get("include") or payload.get("includeKeywords") or payload.get("sports") or payload.get("sport")))
    rules.exclude.extend(split_values(payload.get("exclude") or payload.get("excludeKeywords")))
    for block in split_values(payload.get("blocks") or payload.get("blockRules")):
        rules.blocks.append(parse_rule_line(block, block=True))
    for item in payload.get("ranges") or payload.get("rangeRules") or []:
        if isinstance(item, dict):
            rules.ranges.append(AssignmentRule(
                matcher=str(item.get("matcher") or item.get("sport") or "").strip(),
                min_price=to_number(item.get("min") or item.get("minPrice")),
                max_price=to_number(item.get("max") or item.get("maxPrice")),
            ))
    grades = payload.get("grades") or {}
    if isinstance(grades, dict):
        for company, grade_payload in grades.items():
            if isinstance(grade_payload, dict):
                rules.grade_rules[clean_text(company)] = GradeRule(
                    allowed=grade_payload.get("allowed") is not False,
                    min_grade=to_number(grade_payload.get("min")),
                    max_grade=to_number(grade_payload.get("max")),
                )
    custom_rules = payload.get("rules") or payload.get("customRules") or []
    if isinstance(custom_rules, list):
        rules.rule_groups.extend(parse_custom_rule_group(item) for item in custom_rules if isinstance(item, dict))
        rules.rule_groups = [group for group in rules.rule_groups if group.include or group.ranges or group.grade_rules or group.accept_all]
    return rules


def parse_custom_rule_group(payload: dict[str, Any]) -> CompanyRules:
    group = CompanyRules()
    sports = split_values(payload.get("sports") or payload.get("sport"))
    if payload.get("sportOther"):
        sports.append(str(payload.get("sportOther")).strip())
    group.include.extend(sport for sport in sports if sport and sport != "custom")
    price_ranges = payload.get("priceRanges") or payload.get("ranges") or []
    if isinstance(price_ranges, list):
        for price_range in price_ranges:
            if not isinstance(price_range, dict):
                continue
            min_price = to_number(price_range.get("min") or price_range.get("minPrice"))
            max_price = to_number(price_range.get("max") or price_range.get("maxPrice"))
            if min_price is None and max_price is None:
                continue
            group.ranges.append(AssignmentRule(min_price=min_price, max_price=max_price))
    grades = payload.get("grades") or {}
    if isinstance(grades, dict):
        for company, grade_payload in grades.items():
            if isinstance(grade_payload, dict):
                group.grade_rules[clean_text(company)] = GradeRule(
                    allowed=grade_payload.get("allowed") is not False,
                    min_grade=to_number(grade_payload.get("min")),
                    max_grade=to_number(grade_payload.get("max")),
                )
    return group


def parse_rule_line(line: str, block: bool = False) -> AssignmentRule:
    text = str(line or "").strip()
    over_match = re.match(r"(.+?)\s+(?:over|above)\s+\$?\s*([\d,.]+k?)\+?$", text, re.I)
    if over_match:
        return AssignmentRule(matcher=over_match.group(1).strip(), min_price=parse_money(over_match.group(2)), block=block)
    range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|to|through|thru|–|—)\s*\$?\s*([\d,.]+k?)", text, re.I)
    if range_match:
        matcher = f"{text[:range_match.start()]} {text[range_match.end():]}".strip(" -:|")
        return AssignmentRule(matcher=matcher, min_price=parse_money(range_match.group(1)), max_price=parse_money(range_match.group(2)), block=block)
    return AssignmentRule(matcher=text, block=block)


def parse_payouts(text: str) -> list[PayoutTier]:
    stripped = str(text or "").strip()
    if not stripped:
        return []
    try:
        payload = json.loads(stripped)
        return parse_payout_json(payload)
    except Exception:
        pass

    tiers: list[PayoutTier] = []
    for line in source_lines(stripped):
        rate = parse_rate(line)
        if rate is None:
            continue
        range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|to|through|thru|–|—)\s*\$?\s*([\d,.]+k?)", line, re.I)
        if range_match:
            tiers.append(PayoutTier(parse_money(range_match.group(1)) or 0, parse_money(range_match.group(2)), rate))
            continue
        min_match = re.search(r"(?:over|above|min(?:imum)?)\s+\$?\s*([\d,.]+k?)", line, re.I)
        tiers.append(PayoutTier(parse_money(min_match.group(1)) or 0 if min_match else 0, None, rate))
    return sorted(tiers, key=lambda tier: tier.min_price, reverse=True)


def parse_payout_json(payload: Any) -> list[PayoutTier]:
    if isinstance(payload, dict) and payload.get("rate") is not None:
        rate = parse_rate(payload.get("rate"))
        return [PayoutTier(rate=rate)] if rate is not None else []
    tiers_payload = payload.get("tiers") if isinstance(payload, dict) else payload
    tiers: list[PayoutTier] = []
    for item in tiers_payload or []:
        if not isinstance(item, dict):
            continue
        rate = parse_rate(item.get("rate") or item.get("payout"))
        if rate is None:
            continue
        tiers.append(PayoutTier(
            min_price=to_number(item.get("min") or item.get("minPrice")) or 0,
            max_price=to_number(item.get("max") or item.get("maxPrice")),
            rate=rate,
        ))
    return sorted(tiers, key=lambda tier: tier.min_price, reverse=True)


def source_lines(text: str) -> list[str]:
    rows: list[str] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        try:
            parsed = next(csv.reader([line]))
            line = " ".join(cell.strip() for cell in parsed if cell.strip())
        except Exception:
            pass
        rows.append(line)
    return rows


def rule_matches(rule: AssignmentRule, haystack: str, price: float) -> bool:
    if rule.matcher and not term_matches(rule.matcher, haystack):
        return False
    if rule.min_price is not None and price < rule.min_price:
        return False
    if rule.max_price is not None and price > rule.max_price:
        return False
    return True


def term_matches(term: str, haystack: str) -> bool:
    words = clean_text(term).split()
    if not words:
        return True
    aliases = {
        "b ball": ["basketball", "nba"],
        "bball": ["basketball", "nba"],
        "poke": ["pokemon"],
        "one piece": ["onepiece", "1 piece"],
    }
    options = [words]
    alias_text = " ".join(words)
    options.extend(alias.split() for alias in aliases.get(alias_text, []))
    return any(all(re.search(rf"\b{re.escape(word)}s?\b", haystack) for word in option) for option in options)


def parse_grade(text: str, fallback_grader: str = "") -> tuple[str, float | None]:
    match = re.search(r"\b(PSA|BGS|SGC|CGC)\s*(?:g(?:rade)?\s*)?([0-9]+(?:[._][0-9])?)?\b", text, re.I)
    company = match.group(1).upper() if match else str(fallback_grader or "").upper()
    grade = to_number(match.group(2).replace("_", ".") if match and match.group(2) else None)
    return company, grade


def parse_rate(value: Any) -> float | None:
    text = str(value or "").strip()
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if match:
        return float(match.group(1)) / 100
    number = to_number(text)
    if number is None:
        return None
    return number / 100 if number > 1 else number


def parse_money(value: Any) -> float | None:
    text = str(value or "").strip().lower().replace("$", "").replace(",", "")
    if not text:
        return None
    multiplier = 1000 if text.endswith("k") else 1
    text = text.removesuffix("k")
    return to_number(text) * multiplier if to_number(text) is not None else None


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9/.' -]+", " ", str(value or "").lower())).strip()
