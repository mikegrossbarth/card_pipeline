from __future__ import annotations

import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


DEFAULT_SHEET = "Cards"

PHOTO_EXPORT_HEADERS = {
    "cert": ("certificationnumber", "certnumber", "cert"),
    "description": ("carddescription", "card", "description"),
    "card_number": ("cardnumber",),
    "player": ("playersubject", "player", "subject"),
    "year": ("year",),
    "set": ("set",),
    "subset": ("subset",),
    "parallel": ("parallel",),
    "grader": ("gradingcompany", "grader", "gradingco"),
    "grade": ("grade",),
    "source": ("sourcephoto", "sourcefile"),
}

PHOTO_EXPORT_POSITIONS = {
    "cert": 1,
    "description": 2,
    "card_number": 3,
    "player": 4,
    "year": 5,
    "set": 6,
    "subset": 7,
    "parallel": 8,
    "grader": 9,
    "grade": 10,
    "source": 16,
}


def read_simple_spreadsheet(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows: list[dict[str, Any]] = []
        start_row = 2 if _looks_like_simple_header(sheet) else 1
        for row_index in range(start_row, sheet.max_row + 1):
            cert = normalize_cert(sheet.cell(row_index, 1).value)
            card = clean_part(sheet.cell(row_index, 2).value)
            purchase_price = parse_money(sheet.cell(row_index, 3).value)
            source = clean_part(sheet.cell(row_index, 4).value)
            if not cert and not card and purchase_price is None:
                continue
            rows.append(
                {
                    "cert_number": cert,
                    "card_title": card,
                    "grader": infer_grader(card),
                    "purchase_price": purchase_price,
                    "source": source or f"{path.name}:{row_index}",
                    "notes": _setup_notes(cert, card, infer_grader(card)),
                }
            )
        return rows
    finally:
        workbook.close()


def read_photo_export(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        schema = _detect_photo_schema(sheet)
        rows: list[dict[str, Any]] = []
        for row_index in range(schema["first_data_row"], sheet.max_row + 1):
            source = _source_row(sheet, row_index, schema["headers"])
            cert = normalize_cert(source.get("cert"))
            card = build_card_title(source)
            grader = normalize_grader(source.get("grader")) or infer_grader(card)
            if not cert and not card:
                continue
            rows.append(
                {
                    "cert_number": cert,
                    "card_title": card,
                    "grader": grader,
                    "purchase_price": None,
                    "source": clean_part(source.get("source")) or f"{path.name}:{row_index}",
                    "notes": _setup_notes(cert, card, grader),
                }
            )
        return rows
    finally:
        workbook.close()


def workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def write_pipeline_output(path: Path, rows: list[Any], source_lookup: dict[int, str] | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_SHEET
    headers = [
        "Source",
        "Certification Number",
        "Card Description",
        "Purchase Price",
        "Card Ladder Value",
        "Comps",
        "Card Ladder Comp Details",
        "Card Ladder Screenshot",
        "Best Company",
        "Estimated Payout",
        "Comp Status",
        "Notes",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                (source_lookup or {}).get(row.excel_row, ""),
                row.cert_number,
                row.card_title,
                row.existing_value,
                row.card_ladder_value,
                row.card_ladder_comps_average,
                row.card_ladder_comps,
                row.card_ladder_screenshot,
                row.best_company,
                row.estimated_payout,
                row.status,
                row.notes,
            ]
        )

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [18, 22, 62, 16, 18, 14, 58, 42, 18, 18, 20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(path)
    return path


def write_working_sheet(path: Path, rows: list[Any], source_lookup: dict[int, str] | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_SHEET
    headers = ["Certification Number", "Card Description", "Purchase Price", "Source"]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.cert_number,
            row.card_title,
            row.existing_value,
            (source_lookup or {}).get(row.excel_row, ""),
        ])
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for letter, width in {"A": 22, "B": 62, "C": 16, "D": 38}.items():
        sheet.column_dimensions[letter].width = width
    workbook.save(path)
    return path


def working_sheet_path(directory: Path, title: str) -> Path:
    safe = safe_filename(title) or time.strftime("working-sheet-%Y%m%d-%H%M%S")
    return directory / f"{safe}.xlsx"


def safe_filename(value: str) -> str:
    text = re.sub(r"[<>:\"/\\|?*]+", " ", str(value or "")).strip()
    text = re.sub(r"\s+", " ", text)
    return text[:140].strip()


def default_output_path(root: Path) -> Path:
    stamp = time.strftime("%Y%m%d-%H%M%S")
    return root / "outputs" / f"card-pipeline-comps-{stamp}.xlsx"


def scan_to_cert(value: Any) -> str:
    text = str(value or "")
    candidates = re.findall(r"\d{6,12}", text)
    if candidates:
        return max(candidates, key=len)
    return normalize_cert(text)


def normalize_cert(value: Any) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"[^0-9A-Z]", "", str(value or ""), flags=re.I)


def infer_grader(card_title: str) -> str:
    match = re.search(r"\b(PSA|BGS|SGC|CGC|BECKETT)\b", str(card_title or ""), re.I)
    if not match:
        return ""
    return normalize_grader(match.group(1))


def normalize_grader(value: Any) -> str:
    text = clean_part(value).upper()
    aliases = {"BECKETT": "BGS", "BVG": "BGS", "PSA": "PSA", "BGS": "BGS", "SGC": "SGC", "CGC": "CGC"}
    return aliases.get(text, "")


def parse_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    match = re.search(r"[\d,]+(?:\.\d{1,2})?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def format_money(value: float | None) -> str:
    return "" if value is None else f"${value:,.2f}"


def build_card_title(row: dict[str, Any]) -> str:
    description = clean_part(row.get("description", ""))
    grader = normalize_grader(row.get("grader", ""))
    grade = clean_grade(row.get("grade", ""))
    if description:
        parts = [description]
    else:
        parts = [
            clean_part(row.get("year")),
            clean_part(row.get("set")),
            clean_part(row.get("player")),
            _card_number_part(row.get("card_number")),
            clean_part(row.get("parallel")),
            clean_part(row.get("subset")),
        ]
    if grader and not re.search(rf"\b{re.escape(grader)}\b", " ".join(parts), re.I):
        parts.append(grader)
    if grade and not re.search(rf"(?<!\d){re.escape(grade)}(?!\d)", " ".join(parts)):
        parts.append(grade)
    return re.sub(r"\s+", " ", " ".join(part for part in parts if part)).strip()


def clean_grade(value: Any) -> str:
    numbers = re.findall(r"\d+(?:\.\d+)?", str(value or ""))
    return numbers[-1] if numbers else ""


def clean_part(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _looks_like_simple_header(sheet) -> bool:
    first = " ".join(clean_part(sheet.cell(1, col).value).lower() for col in range(1, min(sheet.max_column, 3) + 1))
    return any(token in first for token in ("cert", "card", "description", "purchase", "price"))


def _setup_notes(cert: str, card: str, grader: str) -> str:
    notes = []
    if not cert:
        notes.append("Missing cert")
    if not card:
        notes.append("Missing card description")
    if not grader:
        notes.append("Missing grader")
    return "; ".join(notes)


def _card_number_part(value: Any) -> str:
    text = clean_part(value)
    if not text:
        return ""
    return text if text.startswith("#") else f"#{text}"


def _detect_photo_schema(sheet) -> dict[str, Any]:
    best_row = None
    best_headers: dict[str, int] = {}
    best_score = 0
    for row_index in range(1, min(sheet.max_row, 10) + 1):
        headers = _header_map_for_row(sheet, row_index)
        score = _header_score(headers)
        if score > best_score:
            best_row = row_index
            best_headers = headers
            best_score = score
    if best_row and best_score >= 3:
        return {"headers": best_headers, "first_data_row": best_row + 1}
    return {"headers": {}, "first_data_row": 1}


def _header_score(headers: dict[str, int]) -> int:
    return sum(1 for aliases in PHOTO_EXPORT_HEADERS.values() if any(alias in headers for alias in aliases))


def _header_map_for_row(sheet, row_index: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row_index, col).value
        if value:
            headers[_normalize_header(value)] = col
    return headers


def _source_row(sheet, row_index: int, headers: dict[str, int]) -> dict[str, Any]:
    return {
        key: _cell(sheet, row_index, headers, aliases, PHOTO_EXPORT_POSITIONS.get(key))
        for key, aliases in PHOTO_EXPORT_HEADERS.items()
    }


def _cell(sheet, row_index: int, headers: dict[str, int], aliases: tuple[str, ...], fallback_col: int | None) -> Any:
    for alias in aliases:
        col = headers.get(alias)
        if col:
            return sheet.cell(row_index, col).value
    return sheet.cell(row_index, fallback_col).value if fallback_col else ""


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())
